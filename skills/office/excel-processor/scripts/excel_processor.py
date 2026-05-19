#!/usr/bin/env python3
"""
Excel 电子表格生成器

用法:
    python excel_processor.py --title "销售报表" --output sales.xlsx
    python excel_processor.py --title "预算表" --output budget.xlsx --sheets data,chart
    python excel_processor.py --dry-run --title "测试"
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart.bar_chart import BarChart
    from openpyxl.chart.reference import Reference
    from openpyxl.chart.series import DataPoint
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 默认配色
THEME = {
    "header_fill": PatternFill(start_color="2C5F8A", end_color="2C5F8A", fill_type="solid"),
    "header_font": Font(color="FFFFFF", bold=True, size=11),
    "title_font": Font(color="2C5F8A", bold=True, size=14),
    "data_font": Font(color="333333", size=10),
    "alt_fill": PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"),
    "thin_border": Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    ),
}


def create_workbook(title="", sheets=None, output_path=None, dry_run=False):
    """创建 Excel 工作簿"""
    if not HAS_OPENPYXL:
        return {"error": "openpyxl 未安装，请运行: pip install openpyxl"}

    if dry_run:
        return {"status": "dry_run", "title": title, "sheets": sheets or [], "message": "仅模拟运行，未生成文件"}

    wb = Workbook()

    sheet_configs = sheets or ["data"]
    created_sheets = []

    for i, sheet_type in enumerate(sheet_configs):
        if i == 0:
            ws = wb.active
            ws.title = f"{title} - 数据"
        else:
            ws = wb.create_sheet(title=f"{title} - {sheet_type}")

        if sheet_type == "data":
            add_data_sheet(ws, title)
            created_sheets.append({"index": i, "type": "data", "title": ws.title})
        elif sheet_type == "chart":
            add_chart_sheet(ws, title)
            created_sheets.append({"index": i, "type": "chart", "title": ws.title})
        elif sheet_type == "summary":
            add_summary_sheet(ws, title)
            created_sheets.append({"index": i, "type": "summary", "title": ws.title})

    if output_path:
        wb.save(output_path)
        return {
            "status": "success",
            "output": str(Path(output_path).resolve()),
            "sheets": len(created_sheets),
            "sheet_details": created_sheets,
        }
    else:
        return {"error": "请指定 --output 参数"}


def add_data_sheet(ws, title):
    """添加数据表"""
    # 标题行
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = THEME["title_font"]
    title_cell.alignment = Alignment(horizontal="left")

    # 表头
    headers = ["日期", "项目", "数量", "单价", "金额"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = THEME["header_font"]
        cell.fill = THEME["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THEME["thin_border"]

    # 示例数据
    sample_data = [
        ["2026-05-01", "产品 A", 100, 50, 5000],
        ["2026-05-02", "产品 B", 200, 30, 6000],
        ["2026-05-03", "产品 C", 150, 40, 6000],
        ["2026-05-04", "产品 A", 80, 50, 4000],
        ["2026-05-05", "产品 B", 120, 30, 3600],
    ]

    for row_idx, row_data in enumerate(sample_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = THEME["data_font"]
            cell.border = THEME["thin_border"]
            cell.alignment = Alignment(horizontal="center")

            # 交替行背景
            if row_idx % 2 == 0:
                cell.fill = THEME["alt_fill"]

            # 金额列格式
            if col_idx == 5:
                cell.number_format = "#,##0"

    # 设置列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12

    # 添加合计行
    total_row = len(sample_data) + 4
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row-1})").font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = "#,##0"


def add_chart_sheet(ws, title):
    """添加图表页"""
    # 标题
    ws.merge_cells("A1:D1")
    ws["A1"].value = f"{title} - 图表分析"
    ws["A1"].font = THEME["title_font"]

    # 数据
    headers = ["月份", "销售额", "成本", "利润"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = THEME["header_font"]
        cell.fill = THEME["header_fill"]
        cell.alignment = Alignment(horizontal="center")

    data = [
        ["1月", 10000, 6000, 4000],
        ["2月", 12000, 7000, 5000],
        ["3月", 15000, 8000, 7000],
        ["4月", 11000, 6500, 4500],
        ["5月", 13000, 7500, 5500],
    ]

    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = THEME["data_font"]
            cell.alignment = Alignment(horizontal="center")
            if col_idx > 1:
                cell.number_format = "#,##0"

    # 创建柱状图
    chart = BarChart()
    chart.type = "col"
    chart.title = "月度销售分析"
    chart.y_axis.title = "金额"
    chart.x_axis.title = "月份"

    data_ref = Reference(ws, min_col=2, min_row=3, max_row=8, max_col=4)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=8)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "A10")


def add_summary_sheet(ws, title):
    """添加汇总页"""
    ws.merge_cells("A1:C1")
    ws["A1"].value = f"{title} - 汇总"
    ws["A1"].font = THEME["title_font"]

    # 汇总指标
    metrics = [
        ["总销售额", "¥51,000"],
        ["总成本", "¥35,000"],
        ["总利润", "¥16,000"],
        ["利润率", "31.4%"],
        ["平均月销售", "¥10,200"],
    ]

    ws.cell(row=3, column=1, value="指标").font = THEME["header_font"]
    ws.cell(row=3, column=1).fill = THEME["header_fill"]
    ws.cell(row=3, column=2, value="数值").font = THEME["header_font"]
    ws.cell(row=3, column=2).fill = THEME["header_fill"]

    for i, (metric, value) in enumerate(metrics, 4):
        ws.cell(row=i, column=1, value=metric).font = THEME["data_font"]
        ws.cell(row=i, column=2, value=value).font = Font(bold=True, size=11)
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = THEME["alt_fill"]
            ws.cell(row=i, column=2).fill = THEME["alt_fill"]

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15


def main():
    parser = argparse.ArgumentParser(description="Excel 电子表格生成器")
    parser.add_argument("--title", default="报表", help="报表标题")
    parser.add_argument("--output", "-o", required=False, help="输出文件路径 (.xlsx)")
    parser.add_argument(
        "--sheets",
        default="data,chart,summary",
        help="工作表类型列表 (data,chart,summary)，逗号分隔",
    )
    parser.add_argument("--dry-run", action="store_true", help="仅模拟运行")
    parser.add_argument("--json", action="store_true", help="输出 JSON 格式")

    args = parser.parse_args()
    sheets = [s.strip() for s in args.sheets.split(",")]

    result = create_workbook(
        title=args.title,
        sheets=sheets,
        output_path=args.output,
        dry_run=args.dry_run,
    )

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        if result.get("error"):
            print(f"❌ {result['error']}", file=sys.stderr)
            sys.exit(1)
        elif result.get("status") == "dry_run":
            print(f"📊 模拟生成: {result['title']}")
            print(f"   工作表: {', '.join(result['sheets'])}")
        else:
            print(f"✅ 已生成: {result['output']}")
            print(f"   共 {result['sheets']} 个工作表")


if __name__ == "__main__":
    main()
